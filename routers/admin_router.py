"""
routers/admin_router.py — Panel de administración: usuarios, metas de equipo.
"""
import io
from datetime import date
from typing import List
from fastapi import APIRouter, Request, Form, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import auth
import database as db

# ── Colores Walmart ──────────────────────────────────────
_BLUE      = "FF0071CE"   # Walmart Blue
_DARK_BLUE = "FF004C99"   # Dark Navy
_YELLOW    = "FFFFC220"   # Spark Yellow
_WHITE     = "FFFFFFFF"
_LIGHT     = "FFF5F8FC"   # fila alterna
_GREEN     = "FF00A651"
_RED       = "FFCC0000"
_ORANGE    = "FFFF6A00"
_AMBER     = "FFFFC107"


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color=None, size=11):
    return Font(bold=bold, color=color or "FF000000", size=size)


def _thin_border() -> Border:
    side = Side(style="thin", color="FFD0D7E2")
    return Border(left=side, right=side, top=side, bottom=side)


def _badge_fill(pct: float) -> str:
    if pct >= 100:
        return _GREEN
    if pct >= 75:
        return _AMBER
    if pct >= 50:
        return _ORANGE
    return _RED


def _build_excel(year: int, month: int,
                 tiendas_filter: list | None = None) -> bytes:
    """Exporta pedidos del mes: Pedido | Determinante | No.Asociado | Tipo."""
    month_name = db.MONTHS_ES[month - 1]
    entries    = db.get_all_entries_for_month(year, month, tiendas_filter=tiendas_filter)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{month_name[:3]} {year}"

    # Titulo
    ws.merge_cells("A1:D1")
    ws["A1"]           = f"Pedidos Omnicanal \u2014 {month_name} {year}"
    ws["A1"].font      = Font(bold=True, color=_WHITE, size=13)
    ws["A1"].fill      = _fill(_DARK_BLUE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Encabezados
    hdrs = ["Pedido", "Determinante", "No. Asociado", "Tipo"]
    for col, h in enumerate(hdrs, 1):
        cell           = ws.cell(row=2, column=col, value=h)
        cell.font      = Font(bold=True, color=_WHITE, size=11)
        cell.fill      = _fill(_BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _thin_border()
    ws.row_dimensions[2].height = 20

    # Filas de datos
    _LIGHT_FILL = PatternFill("solid", fgColor="F0F4FF")
    _WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
    for i, e in enumerate(entries, 1):
        row_num = i + 2
        # Tipo de pedido
        if (e.get("units_kiosko") or 0) > 0:
            tipo = "KIOSCO"
        elif (e.get("units_gana_plus") or 0) > 0:
            tipo = "GANA+"
        else:
            tipo = "OTRO"

        row_data = [
            e.get("order_number") or "",
            e.get("determinante") or "",
            e.get("associate_number") or "",
            tipo,
        ]
        alt = _LIGHT_FILL if i % 2 == 0 else _WHITE_FILL
        for col, val in enumerate(row_data, 1):
            cell           = ws.cell(row=row_num, column=col, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = _thin_border()
            cell.fill      = alt

    # Anchos
    for col, w in zip("ABCD", [22, 14, 16, 10]):
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

router = APIRouter(prefix="/admin")
templates = Jinja2Templates(directory=str(Path(__file__).parent.parent / "templates"))

MONTHS_ES = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
]

SQUADS = ["CENTRO", "NORTE", "OCCIDENTE", "PACIFICO", "SURESTE", "NVDM"]
DISTRITOS = [str(i) for i in range(1, 16)]  # 1 al 15


def _admin_user(request: Request):
    user = auth.get_current_user(request)
    if not user or user["role"] != "admin":
        return None
    return user


def _tiendas_filter(user: dict) -> list | None:
    """Devuelve la lista de tiendas que puede ver este admin.
    None = sin restricción (super admin). Lista = solo esas tiendas.
    """
    tf = db.get_admin_tiendas(user["id"])
    return tf if tf else None


@router.get("", response_class=HTMLResponse)
async def admin_dashboard(request: Request):
    user = _admin_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)

    today = date.today()
    year  = int(request.query_params.get("year",  today.year))
    month = int(request.query_params.get("month", today.month))

    tf        = _tiendas_filter(user)
    team      = db.get_team_productivity(year, month, tiendas_filter=tf)
    matrix    = db.get_team_matrix(year, month, tiendas_filter=tf)
    all_users = db.get_all_users(tiendas_filter=tf)
    wk_start, wk_end = db.commercial_week(today)

    # Tiendas visibles para este admin (para el picker de crear usuario)
    tiendas_disp = tf if tf else db.DETERMINANTES

    top3_asesores = sorted(
        [m for m in team if m["monthly_actual"] > 0],
        key=lambda m: (m["prod_mensual"], m["monthly_actual"]),
        reverse=True,
    )[:3]
    top5 = [m for m in team if m["monthly_actual"] > 0][:5]
    store_budgets_all = db.get_all_store_budgets(year, month)
    # Secundario: filtrar por tiendas asignadas (comparacion normalizada)
    if tf is None:
        store_budgets = store_budgets_all
    else:
        import unicodedata
        def _norm(s: str) -> str:
            return unicodedata.normalize("NFD", s or "").encode("ascii", "ignore").decode().lower().strip()
        tf_norm = {_norm(t) for t in tf}
        store_budgets = [b for b in store_budgets_all if _norm(b["tienda"]) in tf_norm]
    presupuesto_total = sum(b["presupuesto"] for b in store_budgets)
    squad_stats    = db.get_stats_by_group(year, month, "squad")
    distrito_stats = db.get_stats_by_group(year, month, "distrito")
    from tiendas_catalogo import TIENDAS as _TIENDAS_CAT
    import pet_functions as _pets
    import pet_art
    user_ids    = [u["id"] for u in all_users if u["role"] == "user"]
    top3_mascotas = _pets.get_top_pets_for_team(user_ids, year, month, limit=3)
    for i, r in enumerate(top3_mascotas):
        r["rank"]    = i + 1
        r["pet_svg"] = pet_art.get_art(r["animal_type"], r["stage"])
        r["emotion"] = _pets.get_pet_emotion(r)
        r["emeta"]   = _pets.get_emotion_meta(r["emotion"])

    return templates.TemplateResponse("admin.html", {
        "request":          request,
        "user":             user,
        "team":             team,
        "matrix":           matrix,
        "all_users":        all_users,
        "year":             year,
        "month":            month,
        "today":            today.isoformat(),
        "months_es":        MONTHS_ES,
        "determinantes":    tiendas_disp,
        "todas_las_tiendas": db.DETERMINANTES,
        "admin_tiendas":    tf or [],
        "is_super_admin":   tf is None,
        "msg":              request.query_params.get("msg"),
        "error":            request.query_params.get("error"),
        "top5":             top5,
        "top3_asesores":    top3_asesores,
        "top3_mascotas":    top3_mascotas,
        "wk_start":         wk_start.strftime("%d/%m"),
        "wk_end":           wk_end.strftime("%d/%m"),
        "daily_goal":       db.DAILY_GOAL,
        "store_budgets":    store_budgets,
        "presupuesto_total": presupuesto_total,
        "squad_stats":      squad_stats,
        "distrito_stats":   distrito_stats,
        "catalogo_tiendas": _TIENDAS_CAT,
        "squads":           SQUADS,
        "distritos":        DISTRITOS,
    })


@router.post("/users", response_class=HTMLResponse)
async def create_user(
    request: Request,
    associate_number: str = Form(...),
    name: str = Form(...),
    password: str = Form(...),
    role: str = Form("admin_principal"),
    admin_tiendas: List[str] = Form(default=[]),
):
    admin = _admin_user(request)
    if not admin:
        return RedirectResponse("/login", status_code=302)

    if db.associate_number_exists(associate_number):
        return RedirectResponse(
            f"/admin?error=El+numero+{associate_number}+ya+existe", status_code=302
        )

    # Ambos tipos se guardan como role='admin'
    # Principal: sin restriccion de tiendas (ve todo)
    # Secundario: con lista de tiendas asignadas
    es_secundario = (role == "admin_secundario")

    user_id = db.create_user(
        associate_number=associate_number,
        name=name,
        password_hash=auth.hash_password(password),
        role="admin",
        tienda="",
    )
    if es_secundario and admin_tiendas:
        db.set_admin_tiendas(user_id, admin_tiendas)

    return RedirectResponse("/admin?msg=Usuario+admin+creado+correctamente", status_code=302)


@router.post("/team-target", response_class=HTMLResponse)
async def set_team_target(
    request: Request,
    year: int = Form(...),
    month: int = Form(...),
    target_units: float = Form(...),
):
    """Asigna la misma meta a todo el equipo de un solo golpe."""
    if not _admin_user(request):
        return RedirectResponse("/login", status_code=302)

    # Solo aplica a los usuarios visibles para este admin
    admin = _admin_user(request)
    tf    = _tiendas_filter(admin)
    usuarios = db.get_all_users(tiendas_filter=tf)
    for u in usuarios:
        db.set_sales_target(u["id"], year, month, target_units)
    return RedirectResponse(
        f"/admin?year={year}&month={month}&msg=Meta+de+equipo+actualizada", status_code=302
    )


@router.post("/users/{user_id}/delete", response_class=HTMLResponse)
async def delete_user(request: Request, user_id: int):
    """Elimina un usuario y todos sus datos. No se puede borrar a uno mismo."""
    admin = _admin_user(request)
    if not admin:
        return RedirectResponse("/login", status_code=302)
    if admin["id"] == user_id:
        return RedirectResponse("/admin?error=No+puedes+eliminar+tu+propia+cuenta", status_code=302)
    target = db.get_user_by_id(user_id)
    if not target:
        return RedirectResponse("/admin?error=Usuario+no+encontrado", status_code=302)
    db.delete_user(user_id)
    nombre = target["name"]
    return RedirectResponse(f"/admin?msg=Usuario+{nombre}+eliminado+correctamente", status_code=302)


@router.post("/reset-password", response_class=HTMLResponse)
async def reset_password(
    request: Request,
    user_id: int = Form(...),
    new_password: str = Form(...),
):
    if not _admin_user(request):
        return RedirectResponse("/login", status_code=302)

    # Actualiza la contrasena directamente — el asociado entra sin friccion
    db.update_user_password(user_id, auth.hash_password(new_password), clear_force=True)
    return RedirectResponse("/admin?msg=Contrasena+restablecida+correctamente", status_code=302)


@router.get("/export/excel")
async def export_excel(request: Request):
    """Descarga el reporte del equipo en formato .xlsx."""
    user = _admin_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)

    today = date.today()
    year  = int(request.query_params.get("year",  today.year))
    month = int(request.query_params.get("month", today.month))
    month_name = MONTHS_ES[month - 1]

    tf         = _tiendas_filter(user)
    xlsx_bytes = _build_excel(year, month, tiendas_filter=tf)
    filename   = f"productividad_{month_name.lower()}_{year}.xlsx"

    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── PRESUPUESTOS POR TIENDA ─────────────────────────────────────────────────

@router.get("/presupuestos", response_class=HTMLResponse)
async def page_presupuestos(request: Request):
    """Pagina dedicada a gestion de presupuestos por tienda."""
    user = _admin_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)

    today = date.today()
    year  = int(request.query_params.get("year",  today.year))
    month = int(request.query_params.get("month", today.month))

    # Presupuestos mensuales del periodo
    raw_budgets = db.get_all_store_budgets(year, month)

    # Enriquecer con conteo de dias y suma de diarios
    store_budgets = []
    for sb in raw_budgets:
        det  = sb["determinante"]
        daily = db.get_store_daily_budgets_month(det, year, month)
        suma  = sum(daily.values())
        diff  = sb["presupuesto"] - suma
        store_budgets.append({
            **sb,
            "dias_cargados": len(daily),
            "suma_diarios":  round(suma, 0),
            "diferencia":    round(diff, 0),
        })

    return templates.TemplateResponse("admin_presupuestos.html", {
        "request":      request,
        "user":         user,
        "year":         year,
        "month":        month,
        "months_es":    MONTHS_ES,
        "store_budgets": store_budgets,
        "msg":          request.query_params.get("msg"),
        "error":        request.query_params.get("error"),
    })


@router.get("/presupuesto/plantilla")
async def descargar_plantilla_presupuesto(request: Request):
    """Excel con 1071 tiendas + columnas de presupuesto mensual y dias del mes."""
    user = _admin_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)

    import calendar
    today  = date.today()
    year   = int(request.query_params.get("year",  today.year))
    month  = int(request.query_params.get("month", today.month))
    m_name = MONTHS_ES[month - 1]
    n_days = calendar.monthrange(year, month)[1]

    from tiendas_catalogo import TIENDAS
    saved_mensual = {r["determinante"]: r["presupuesto"]
                    for r in db.get_all_store_budgets(year, month)}

    wb = Workbook()
    ws = wb.active
    ws.title = "Presupuestos"

    hdrs = ["Determinante", "Tienda", "Presupuesto_Mensual_Omnicanal"] + list(range(1, n_days + 1))
    for col, h in enumerate(hdrs, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = Font(bold=True, color=_WHITE, size=10)
        cell.fill      = _fill(_BLUE)
        cell.alignment = Alignment(horizontal="center")

    for row_i, (det, nombre) in enumerate(TIENDAS, 2):
        ws.cell(row=row_i, column=1, value=det)
        ws.cell(row=row_i, column=2, value=nombre)
        ws.cell(row=row_i, column=3, value=saved_mensual.get(det, ""))

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 26
    for col_i in range(4, 4 + n_days):
        ws.column_dimensions[get_column_letter(col_i)].width = 9

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"presupuestos_{m_name.lower()}_{year}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/presupuesto/subir", response_class=HTMLResponse)
async def subir_presupuesto(request: Request, archivo: UploadFile = File(...)):
    """
    Lee Excel con formato:
      Determinante | Tienda | Presupuesto_Mensual | Dia 1 | Dia 2 | ... | Dia 31
    Guarda presupuesto mensual + diarios en DB.
    """
    user = _admin_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)

    today = date.today()
    try:
        contenido = await archivo.read()
        wb_in = load_workbook(filename=io.BytesIO(contenido), read_only=True, data_only=True)
        ws_in = wb_in.active

        # ── Leer encabezado (fila 1) ─────────────────────────────────────────
        header_row = next(ws_in.iter_rows(min_row=1, max_row=1, values_only=True), [])
        headers = [str(h).lower().strip() if h is not None else "" for h in header_row]

        def _find_col(*keywords):
            for idx, h in enumerate(headers):
                if any(kw in h for kw in keywords):
                    return idx
            return None

        col_det   = _find_col("determinante", "det", "clave", "id")
        col_tiend = _find_col("tienda", "store", "sucursal", "nombre")
        col_pres  = _find_col("mensual", "presupuesto", "omnicanal", "meta", "budget", "importe")
        col_year  = _find_col("a\u00f1o", "anio", "year")
        col_month = _find_col("mes", "month")

        # Detectar columnas de dias: encabezados que sean numeros 1-31
        # o que contengan 'dia', 'd1', 'day', etc.
        day_cols: dict[int, int] = {}  # {numero_dia: indice_columna}
        for idx, h in enumerate(headers):
            # Header numerico puro (el dia)
            if h.isdigit() and 1 <= int(h) <= 31:
                day_cols[int(h)] = idx
                continue
            # Header tipo "dia 5", "d5", "day5"
            import re
            m = re.match(r'^(?:d[\u00eda]*a?|day)\s*(\d{1,2})$', h)
            if m and 1 <= int(m.group(1)) <= 31:
                day_cols[int(m.group(1))] = idx

        # Fallback posicional si no hay headers utiles
        if col_det is None:   col_det = 0
        if col_tiend is None: col_tiend = 1
        if col_pres is None:  col_pres = 2
        # Si no hay columnas de dia detectadas por nombre, asumir col 3 en adelante = dia 1,2,3...
        if not day_cols:
            for extra_idx in range(3, min(len(headers), 34)):
                dia = extra_idx - 2  # col 3 -> dia 1, col 4 -> dia 2...
                day_cols[dia] = extra_idx

        rows_ok: list[dict] = []
        errores: list[str]  = []

        for i, row in enumerate(ws_in.iter_rows(min_row=2, values_only=True), 2):
            if not row or len(row) <= col_det:
                continue
            if row[col_det] is None:
                continue

            det   = str(row[col_det]).strip().lstrip('0') or str(row[col_det]).strip()
            tiend = str(row[col_tiend]).strip() if col_tiend < len(row) and row[col_tiend] else ""

            # Presupuesto mensual
            raw_pres = row[col_pres] if col_pres < len(row) else None
            if raw_pres is None:
                errores.append(f"Fila {i}: sin presupuesto mensual para det '{det}'")
                continue
            try:
                clean = str(raw_pres).replace("$", "").replace(",", "").strip()
                presupuesto = float(clean)
            except (ValueError, TypeError):
                errores.append(f"Fila {i}: presupuesto '{raw_pres}' no es numero")
                continue

            if presupuesto < 0:
                errores.append(f"Fila {i}: presupuesto negativo para '{det}'")
                continue

            # Ano y mes
            try:
                year_val  = int(row[col_year])  if col_year  and col_year  < len(row) and row[col_year]  else today.year
                month_val = int(row[col_month]) if col_month and col_month < len(row) and row[col_month] else today.month
            except (ValueError, TypeError):
                year_val, month_val = today.year, today.month

            # Presupuestos diarios
            dias: dict[int, float] = {}
            for day_num, day_col in day_cols.items():
                if day_col >= len(row) or row[day_col] is None:
                    continue
                try:
                    val_dia = float(str(row[day_col]).replace("$", "").replace(",", "").strip())
                    if val_dia >= 0:
                        dias[day_num] = val_dia
                except (ValueError, TypeError):
                    pass  # Celda vacia o texto -> ignorar

            rows_ok.append({
                "determinante": det,
                "tienda":       tiend,
                "year":         year_val,
                "month":        month_val,
                "presupuesto":  presupuesto,
                "dias":         dias,
            })

        if not rows_ok:
            cols_info = "|".join(headers[:8]) or "ninguna"
            return RedirectResponse(
                f"/admin?error=Sin+filas+validas.+Columnas+detectadas:+{cols_info}",
                status_code=302,
            )

        n   = db.upsert_store_budgets(rows_ok)
        total_dias = sum(len(r["dias"]) for r in rows_ok)
        msg = f"{n}+tienda(s)+cargadas+con+{total_dias}+presupuestos+diarios"
        if errores:
            msg += f".+{len(errores)}+fila(s)+con+problema"
        return RedirectResponse(f"/admin/presupuestos?msg={msg}", status_code=302)

    except Exception as exc:
        return RedirectResponse(
            f"/admin/presupuestos?error=Error+leyendo+archivo:+{str(exc)[:100]}", status_code=302
        )
